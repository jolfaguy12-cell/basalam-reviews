import urllib.request
import json
import csv
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

VENDOR_ID = 1399163
VENDOR_IDENTIFIER = "behdashtik"
BASE_URL = f"https://services.basalam.com/web/v1/review/vendor/{VENDOR_ID}/reviews"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
LIMIT = 20


def fetch_all_reviews():
    all_reviews = []
    offset = 0
    page = 1

    while True:
        url = f"{BASE_URL}?limit={LIMIT}&offset={offset}"
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req) as resp:
            data = json.load(resp)

        reviews = data.get("reviews", [])
        all_reviews.extend(reviews)
        print(f"Page {page}: fetched {len(reviews)} (total: {len(all_reviews)})")

        if not data.get("has_next") or not reviews:
            break

        offset += LIMIT
        page += 1
        time.sleep(0.3)

    return all_reviews


def flatten(reviews):
    rows = []
    for r in reviews:
        answers = r.get("answers", [])
        reply_text = " | ".join(a.get("description", "") for a in answers)
        reply_by   = " | ".join(a.get("user", {}).get("name", "") for a in answers)
        rows.append({
            "review_id":  r["id"],
            "date":       r["createdAt"][:10],
            "user_name":  r["user"]["name"],
            "stars":      r["star"],
            "review":     r.get("description", ""),
            "product_id": r.get("productId", ""),
            "product":    r.get("product", {}).get("title", ""),
            "reply_by":   reply_by,
            "reply":      reply_text,
        })
    return rows


def save_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"CSV saved: {path}")


def save_excel(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reviews"

    headers = ["Review ID", "Date", "User", "Stars", "Review",
               "Product ID", "Product", "Reply By", "Reply"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 12, 20, 7, 60, 12, 60, 20, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    star_fills = {
        5: PatternFill("solid", fgColor="E2EFDA"),
        4: PatternFill("solid", fgColor="EBF3DE"),
        3: PatternFill("solid", fgColor="FFF2CC"),
        2: PatternFill("solid", fgColor="FCE4D6"),
        1: PatternFill("solid", fgColor="F4CCCC"),
    }

    for row_idx, r in enumerate(rows, 2):
        vals = [r["review_id"], r["date"], r["user_name"], r["stars"],
                r["review"], r["product_id"], r["product"], r["reply_by"], r["reply"]]
        fill = star_fills.get(r["stars"])
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"Excel saved: {path}")


if __name__ == "__main__":
    reviews = fetch_all_reviews()
    rows = flatten(reviews)
    save_csv(rows, f"{VENDOR_IDENTIFIER}_reviews.csv")
    save_excel(rows, f"{VENDOR_IDENTIFIER}_reviews.xlsx")
    print(f"\nDone — {len(rows)} reviews exported.")
